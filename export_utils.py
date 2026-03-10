"""
Export module - generates PDF and Excel reports
Uses only standard library + reportlab + openpyxl
"""

import os
import csv
from datetime import datetime

# ─── Try importing optional libraries ─────────────────────────────────────────

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def get_export_dir():
    if os.name == 'nt':
        path = os.path.join(os.path.expanduser('~'), 'Documents', 'ConstructionPro_Exports')
    else:
        path = os.path.join(os.path.expanduser('~'), 'ConstructionPro_Exports')
    os.makedirs(path, exist_ok=True)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORTS
# ══════════════════════════════════════════════════════════════════════════════

def _pdf_header_style():
    styles = getSampleStyleSheet()
    return {
        'title': ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=16,
                                 textColor=colors.HexColor('#1a3c5e'), alignment=TA_CENTER),
        'subtitle': ParagraphStyle('Sub', fontName='Helvetica', fontSize=10,
                                    textColor=colors.HexColor('#555555'), alignment=TA_CENTER),
        'heading': ParagraphStyle('Heading', fontName='Helvetica-Bold', fontSize=12,
                                   textColor=colors.HexColor('#1a3c5e')),
        'normal': ParagraphStyle('Normal', fontName='Helvetica', fontSize=9,
                                  textColor=colors.black),
        'right': ParagraphStyle('Right', fontName='Helvetica', fontSize=9,
                                 alignment=TA_RIGHT),
    }

def _table_style_default():
    return TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a3c5e')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ])


def export_employee_pdf(employee, attendance, fuel_logs, settings):
    if not REPORTLAB_OK:
        return None, "ReportLab not installed. Run: pip install reportlab"

    fname = f"Employee_{employee['name'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    fpath = os.path.join(get_export_dir(), fname)
    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    s = _pdf_header_style()
    story = []

    # Header
    story.append(Paragraph(settings.get('company_name', 'Construction Co.'), s['title']))
    story.append(Paragraph(settings.get('company_address', ''), s['subtitle']))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a3c5e')))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Employee Report: {employee['name']}", s['heading']))
    story.append(Spacer(1, 8))

    # Employee details
    details = [
        ['Field', 'Details'],
        ['Name', employee['name']],
        ['Role', employee['role']],
        ['Phone', employee.get('phone', '-')],
        ['Address', employee.get('address', '-')],
        ['Daily Wage', f"Rs. {employee.get('daily_wage', 0):.2f}"],
        ['Joining Date', employee.get('joining_date', '-')],
    ]
    t = Table(details, colWidths=[5*cm, 12*cm])
    t.setStyle(_table_style_default())
    story.append(t)
    story.append(Spacer(1, 15))

    # Attendance
    story.append(Paragraph("Attendance Records", s['heading']))
    story.append(Spacer(1, 5))
    if attendance:
        present = sum(1 for a in attendance if a['status'] == 'present')
        absent = sum(1 for a in attendance if a['status'] == 'absent')
        half = sum(1 for a in attendance if a['status'] == 'half_day')
        story.append(Paragraph(
            f"Total Days: {len(attendance)} | Present: {present} | Absent: {absent} | Half Day: {half}",
            s['normal']
        ))
        story.append(Spacer(1, 5))
        att_data = [['Date', 'Status', 'Notes']]
        for a in attendance:
            att_data.append([a['date'], a['status'].replace('_', ' ').title(), a.get('notes', '') or ''])
        t2 = Table(att_data, colWidths=[4*cm, 4*cm, 9*cm])
        t2.setStyle(_table_style_default())
        story.append(t2)
    else:
        story.append(Paragraph("No attendance records found.", s['normal']))

    # Fuel logs (drivers)
    if fuel_logs:
        story.append(Spacer(1, 15))
        story.append(Paragraph("Diesel / Fuel Log", s['heading']))
        story.append(Spacer(1, 5))
        total_liters = sum(f['liters'] for f in fuel_logs)
        total_amount = sum(f['amount'] for f in fuel_logs)
        story.append(Paragraph(
            f"Total Fuel: {total_liters:.2f} L | Total Amount: Rs. {total_amount:.2f}",
            s['normal']
        ))
        story.append(Spacer(1, 5))
        fuel_data = [['Date', 'Vehicle', 'Liters', 'Amount (Rs.)', 'Notes']]
        for f in fuel_logs:
            fuel_data.append([f['date'], f.get('vehicle', '-'),
                               f"{f['liters']:.2f}", f"Rs.{f['amount']:.2f}",
                               f.get('notes', '') or ''])
        t3 = Table(fuel_data, colWidths=[3.5*cm, 3.5*cm, 2.5*cm, 3.5*cm, 4*cm])
        t3.setStyle(_table_style_default())
        story.append(t3)

    story.append(Spacer(1, 15))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}", s['subtitle']))

    doc.build(story)
    return fpath, None


def export_materials_pdf(materials, settings, from_date='', to_date=''):
    if not REPORTLAB_OK:
        return None, "ReportLab not installed."

    fname = f"Materials_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    fpath = os.path.join(get_export_dir(), fname)
    doc = SimpleDocTemplate(fpath, pagesize=landscape(A4),
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    s = _pdf_header_style()
    story = []

    story.append(Paragraph(settings.get('company_name', 'Construction Co.'), s['title']))
    story.append(Paragraph("Material Register", s['subtitle']))
    if from_date or to_date:
        story.append(Paragraph(f"Period: {from_date or 'All'} to {to_date or 'All'}", s['subtitle']))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a3c5e')))
    story.append(Spacer(1, 10))

    if materials:
        total_net = sum(m.get('net_weight', 0) for m in materials)
        total_amt = sum(m.get('amount', 0) for m in materials)
        story.append(Paragraph(
            f"Total Records: {len(materials)} | Total Net Weight: {total_net:.2f} | Total Amount: Rs. {total_amt:.2f}",
            s['normal']
        ))
        story.append(Spacer(1, 8))
        mat_data = [['Date', 'Material', 'Load Wt', 'Empty Wt', 'Net Wt', 'Supplier', 'Vehicle', 'Rate', 'Amount']]
        for m in materials:
            mat_data.append([
                m['date'], m['material_name'],
                f"{m.get('load_weight',0):.2f}", f"{m.get('empty_weight',0):.2f}",
                f"{m.get('net_weight',0):.2f}", m.get('supplier', '-'),
                m.get('vehicle_no', '-'), f"Rs.{m.get('rate',0):.2f}",
                f"Rs.{m.get('amount',0):.2f}"
            ])
        # Summary row
        mat_data.append(['', 'TOTAL', '', '', f"{total_net:.2f}", '', '', '', f"Rs.{total_amt:.2f}"])

        col_widths = [3*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 4*cm, 3*cm, 2.5*cm, 3*cm]
        t = Table(mat_data, colWidths=col_widths)
        style = _table_style_default()
        # Bold last row
        style.add('FONTNAME', (0, len(mat_data)-1), (-1, len(mat_data)-1), 'Helvetica-Bold')
        style.add('BACKGROUND', (0, len(mat_data)-1), (-1, len(mat_data)-1), colors.HexColor('#d4e8f0'))
        t.setStyle(style)
        story.append(t)
    else:
        story.append(Paragraph("No material records found.", s['normal']))

    story.append(Spacer(1, 15))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}", s['subtitle']))
    doc.build(story)
    return fpath, None


def export_gst_bill_pdf(bill, items, settings):
    if not REPORTLAB_OK:
        return None, "ReportLab not installed."

    fname = f"GST_{bill['bill_no']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    fpath = os.path.join(get_export_dir(), fname)
    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    s = _pdf_header_style()
    story = []

    # Company header
    story.append(Paragraph(settings.get('company_name', 'Construction Co.'), s['title']))
    story.append(Paragraph(settings.get('company_address', ''), s['subtitle']))
    story.append(Paragraph(f"GSTIN: {settings.get('company_gstin', '')}  |  Ph: {settings.get('company_phone', '')}", s['subtitle']))
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a3c5e')))
    story.append(Spacer(1, 6))
    story.append(Paragraph("TAX INVOICE", ParagraphStyle('TI', fontName='Helvetica-Bold',
                                                           fontSize=14, alignment=TA_CENTER,
                                                           textColor=colors.HexColor('#1a3c5e'))))
    story.append(Spacer(1, 8))

    # Bill & Client info
    bill_info = [
        ['Bill No:', bill['bill_no'], 'Bill Date:', bill['bill_date']],
        ['Client:', bill['client_name'], 'GSTIN:', bill.get('client_gstin', '-')],
        ['Address:', bill.get('client_address', '-'), '', ''],
    ]
    t_info = Table(bill_info, colWidths=[3*cm, 8*cm, 3*cm, 5*cm])
    t_info.setStyle(TableStyle([
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_info)
    story.append(Spacer(1, 12))

    # Items table
    item_data = [['#', 'Description', 'HSN', 'Qty', 'Unit', 'Rate (Rs.)', 'Amount (Rs.)']]
    for i, item in enumerate(items, 1):
        item_data.append([
            str(i), item['description'], item.get('hsn_code', '-'),
            f"{item['quantity']}", item.get('unit', 'Nos'),
            f"Rs.{item['rate']:.2f}", f"Rs.{item['amount']:.2f}"
        ])

    t_items = Table(item_data, colWidths=[1*cm, 6*cm, 2*cm, 1.5*cm, 1.5*cm, 3*cm, 3*cm])
    t_items.setStyle(_table_style_default())
    story.append(t_items)
    story.append(Spacer(1, 10))

    # Totals
    subtotal = bill['subtotal']
    cgst = bill['cgst_amount']
    sgst = bill['sgst_amount']
    total = bill['total']

    totals_data = [
        ['', '', '', '', '', 'Sub Total:', f"Rs. {subtotal:.2f}"],
        ['', '', '', '', '', f"CGST @ {bill['cgst_rate']}%:", f"Rs. {cgst:.2f}"],
        ['', '', '', '', '', f"SGST @ {bill['sgst_rate']}%:", f"Rs. {sgst:.2f}"],
        ['', '', '', '', '', 'TOTAL:', f"Rs. {total:.2f}"],
    ]
    t_totals = Table(totals_data, colWidths=[1*cm, 6*cm, 2*cm, 1.5*cm, 1.5*cm, 3*cm, 3*cm])
    t_totals.setStyle(TableStyle([
        ('FONTNAME', (5,0), (6,-1), 'Helvetica'),
        ('FONTNAME', (5,3), (6,3), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (5,0), (6,-1), 'RIGHT'),
        ('BACKGROUND', (5,3), (6,3), colors.HexColor('#1a3c5e')),
        ('TEXTCOLOR', (5,3), (6,3), colors.white),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('GRID', (5,0), (6,-1), 0.5, colors.HexColor('#cccccc')),
    ]))
    story.append(t_totals)
    story.append(Spacer(1, 20))

    # Signature area
    sig_data = [['Authorised Signatory', '']]
    t_sig = Table(sig_data, colWidths=[8*cm, 9*cm])
    t_sig.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (0,0), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 30),
        ('BOX', (0,0), (0,0), 0.5, colors.grey),
    ]))
    story.append(t_sig)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}", s['subtitle']))

    doc.build(story)
    return fpath, None


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORTS
# ══════════════════════════════════════════════════════════════════════════════

def _excel_header_style(ws, company_name):
    """Apply common header to worksheet"""
    if not OPENPYXL_OK:
        return
    ws['A1'] = company_name
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1A3C5E')
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')

def _hdr_fill():
    return PatternFill('solid', fgColor='1A3C5E')

def _hdr_font():
    return Font(bold=True, color='FFFFFF', name='Arial', size=10)

def _border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_employee_excel(employee, attendance, fuel_logs, settings):
    if not OPENPYXL_OK:
        return None, "openpyxl not installed. Run: pip install openpyxl"

    fname = f"Employee_{employee['name'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    fpath = os.path.join(get_export_dir(), fname)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Employee Info ──
    ws1 = wb.active
    ws1.title = 'Employee Info'
    _excel_header_style(ws1, settings.get('company_name', 'Construction Co.'))

    ws1.append([])
    ws1.append(['Employee Details'])
    ws1['A3'].font = Font(bold=True, size=12, color='1A3C5E')

    fields = [
        ('Name', employee['name']),
        ('Role', employee['role']),
        ('Phone', employee.get('phone', '-')),
        ('Address', employee.get('address', '-')),
        ('Daily Wage', f"Rs. {employee.get('daily_wage', 0):.2f}"),
        ('Joining Date', employee.get('joining_date', '-')),
    ]
    for field, value in fields:
        ws1.append([field, value])
        row = ws1.max_row
        ws1.cell(row, 1).font = Font(bold=True, name='Arial')
        ws1.cell(row, 2).font = Font(name='Arial')

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 40

    # ── Sheet 2: Attendance ──
    ws2 = wb.create_sheet('Attendance')
    _excel_header_style(ws2, settings.get('company_name', ''))
    ws2.append([])
    headers = ['Date', 'Status', 'Notes']
    ws2.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(ws2.max_row, col)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
        cell.border = _border()
        cell.alignment = Alignment(horizontal='center')

    for a in attendance:
        ws2.append([a['date'], a['status'].replace('_', ' ').title(), a.get('notes', '') or ''])
        for col in range(1, 4):
            ws2.cell(ws2.max_row, col).border = _border()

    # Summary
    ws2.append([])
    present = sum(1 for a in attendance if a['status'] == 'present')
    absent = sum(1 for a in attendance if a['status'] == 'absent')
    half = sum(1 for a in attendance if a['status'] == 'half_day')
    ws2.append(['Summary:', f'Total: {len(attendance)}', f'Present: {present}', f'Absent: {absent}', f'Half Day: {half}'])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True)

    for col_letter, width in [('A', 15), ('B', 18), ('C', 40)]:
        ws2.column_dimensions[col_letter].width = width

    # ── Sheet 3: Fuel Log ──
    if fuel_logs:
        ws3 = wb.create_sheet('Fuel Log')
        _excel_header_style(ws3, settings.get('company_name', ''))
        ws3.append([])
        headers3 = ['Date', 'Vehicle', 'Liters', 'Amount (Rs.)', 'Notes']
        ws3.append(headers3)
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(ws3.max_row, col)
            cell.fill = _hdr_fill()
            cell.font = _hdr_font()
            cell.border = _border()
            cell.alignment = Alignment(horizontal='center')

        for f in fuel_logs:
            ws3.append([f['date'], f.get('vehicle', '-'),
                        f['liters'], f['amount'], f.get('notes', '') or ''])
            for col in range(1, 6):
                ws3.cell(ws3.max_row, col).border = _border()

        # Totals
        ws3.append(['', 'TOTAL',
                    sum(f['liters'] for f in fuel_logs),
                    sum(f['amount'] for f in fuel_logs), ''])
        for col in [1, 2, 3, 4, 5]:
            ws3.cell(ws3.max_row, col).font = Font(bold=True)

        for col_letter, width in [('A', 15), ('B', 20), ('C', 12), ('D', 15), ('E', 30)]:
            ws3.column_dimensions[col_letter].width = width

    wb.save(fpath)
    return fpath, None


def export_materials_excel(materials, settings, from_date='', to_date=''):
    if not OPENPYXL_OK:
        return None, "openpyxl not installed."

    fname = f"Materials_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    fpath = os.path.join(get_export_dir(), fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materials'

    _excel_header_style(ws, settings.get('company_name', 'Construction Co.'))
    ws.append([])
    ws.append([f'Material Register{" | Period: "+from_date+" to "+to_date if from_date else ""}'])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color='1A3C5E')
    ws.append([])

    headers = ['Date', 'Material Name', 'Load Weight', 'Empty Weight', 'Net Weight',
               'Supplier', 'Vehicle No', 'Rate (Rs.)', 'Amount (Rs.)', 'Notes']
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(ws.max_row, col)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
        cell.border = _border()
        cell.alignment = Alignment(horizontal='center')

    even_fill = PatternFill('solid', fgColor='EEF4F8')
    for i, m in enumerate(materials):
        row = [m['date'], m['material_name'],
               m.get('load_weight', 0), m.get('empty_weight', 0), m.get('net_weight', 0),
               m.get('supplier', '-'), m.get('vehicle_no', '-'),
               m.get('rate', 0), m.get('amount', 0), m.get('notes', '') or '']
        ws.append(row)
        fill = even_fill if i % 2 == 1 else None
        for col in range(1, 11):
            cell = ws.cell(ws.max_row, col)
            cell.border = _border()
            if fill:
                cell.fill = fill

    # Totals row
    total_net = sum(m.get('net_weight', 0) for m in materials)
    total_amt = sum(m.get('amount', 0) for m in materials)
    ws.append(['', 'TOTAL', '', '', total_net, '', '', '', total_amt, ''])
    for col in range(1, 11):
        cell = ws.cell(ws.max_row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D4E8F0')
        cell.border = _border()

    widths = [('A', 15), ('B', 25), ('C', 14), ('D', 14), ('E', 14),
              ('F', 25), ('G', 15), ('H', 12), ('I', 14), ('J', 30)]
    for col_letter, width in widths:
        ws.column_dimensions[col_letter].width = width

    wb.save(fpath)
    return fpath, None


def export_gst_bill_excel(bill, items, settings):
    if not OPENPYXL_OK:
        return None, "openpyxl not installed."

    fname = f"GST_{bill['bill_no']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    fpath = os.path.join(get_export_dir(), fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tax Invoice'

    # Company header
    ws['A1'] = settings.get('company_name', 'Construction Co.')
    ws['A1'].font = Font(bold=True, size=16, color='1A3C5E')
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = settings.get('company_address', '')
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=9, color='555555')

    ws['A3'] = f"GSTIN: {settings.get('company_gstin', '')}  |  Phone: {settings.get('company_phone', '')}"
    ws.merge_cells('A3:G3')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A3'].font = Font(size=9, color='555555')

    ws.append([])
    ws['A5'] = 'TAX INVOICE'
    ws['A5'].font = Font(bold=True, size=14, color='1A3C5E')
    ws.merge_cells('A5:G5')
    ws['A5'].alignment = Alignment(horizontal='center')
    ws.append([])

    # Bill info
    ws.append(['Bill No:', bill['bill_no'], '', 'Bill Date:', bill['bill_date']])
    ws.append(['Client:', bill['client_name'], '', 'GSTIN:', bill.get('client_gstin', '-')])
    ws.append(['Address:', bill.get('client_address', '-')])
    ws.append([])

    # Items
    headers = ['#', 'Description', 'HSN Code', 'Qty', 'Unit', 'Rate (Rs.)', 'Amount (Rs.)']
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(ws.max_row, col)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
        cell.border = _border()
        cell.alignment = Alignment(horizontal='center')

    for i, item in enumerate(items, 1):
        ws.append([i, item['description'], item.get('hsn_code', '-'),
                   item['quantity'], item.get('unit', 'Nos'),
                   item['rate'], item['amount']])
        for col in range(1, 8):
            ws.cell(ws.max_row, col).border = _border()

    ws.append([])

    # Totals
    for label, value in [
        ('Sub Total', f"Rs. {bill['subtotal']:.2f}"),
        (f"CGST @ {bill['cgst_rate']}%", f"Rs. {bill['cgst_amount']:.2f}"),
        (f"SGST @ {bill['sgst_rate']}%", f"Rs. {bill['sgst_amount']:.2f}"),
        ('GRAND TOTAL', f"Rs. {bill['total']:.2f}"),
    ]:
        ws.append(['', '', '', '', '', label, value])
        ws.cell(ws.max_row, 6).font = Font(bold=True)
        ws.cell(ws.max_row, 7).font = Font(bold=True)
        if label == 'GRAND TOTAL':
            ws.cell(ws.max_row, 6).fill = PatternFill('solid', fgColor='1A3C5E')
            ws.cell(ws.max_row, 7).fill = PatternFill('solid', fgColor='1A3C5E')
            ws.cell(ws.max_row, 6).font = Font(bold=True, color='FFFFFF')
            ws.cell(ws.max_row, 7).font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    wb.save(fpath)
    return fpath, None


# ══════════════════════════════════════════════════════════════════════════════
#  DIESEL FUEL LOG EXPORTS
# ══════════════════════════════════════════════════════════════════════════════

def export_diesel_pdf(records, settings, from_date='', to_date=''):
    if not REPORTLAB_OK:
        return None, "ReportLab not installed. Run: pip install reportlab"
    fname = f"DieselFuelLog_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    fpath = os.path.join(get_export_dir(), fname)
    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    s = _pdf_header_style()
    story = []

    story.append(Paragraph(settings.get('company_name', 'Construction Co.'), s['title']))
    story.append(Paragraph('Diesel Fuel Details Log', s['subtitle']))
    if from_date or to_date:
        story.append(Paragraph(f"Period: {from_date or 'All'} to {to_date or 'All'}", s['subtitle']))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a3c5e')))
    story.append(Spacer(1, 10))

    if records:
        total_qty = sum(r.get('qty_liters', 0) for r in records)
        total_amt = sum(r.get('amount', 0) for r in records)
        story.append(Paragraph(
            f"Total Entries: {len(records)}  |  Total Qty: {total_qty:.2f} Ltrs  |  Total Amount: Rs. {total_amt:,.2f}",
            s['normal']
        ))
        story.append(Spacer(1, 8))

        data = [['I.No', 'Date', 'Vehicle No.', 'Qty (Ltrs)', 'Amount (Rs.)']]
        for r in records:
            data.append([
                str(r['sl_no']), r['date'], r['vehicle_no'],
                f"{r['qty_liters']:.2f}", f"Rs. {r['amount']:.2f}"
            ])
        # Totals row
        data.append(['', 'TOTAL', '', f"{total_qty:.2f}", f"Rs. {total_amt:,.2f}"])

        t = Table(data, colWidths=[2*cm, 4*cm, 5*cm, 4*cm, 4*cm])
        style = _table_style_default()
        style.add('FONTNAME', (0, len(data)-1), (-1, len(data)-1), 'Helvetica-Bold')
        style.add('BACKGROUND', (0, len(data)-1), (-1, len(data)-1), colors.HexColor('#d4e8f0'))
        t.setStyle(style)
        story.append(t)
    else:
        story.append(Paragraph("No diesel fuel records found.", s['normal']))

    story.append(Spacer(1, 15))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}", s['subtitle']))
    doc.build(story)
    return fpath, None


def export_diesel_excel(records, settings, from_date='', to_date=''):
    if not OPENPYXL_OK:
        return None, "openpyxl not installed. Run: pip install openpyxl"
    fname = f"DieselFuelLog_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    fpath = os.path.join(get_export_dir(), fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Diesel Fuel Log'

    _excel_header_style(ws, settings.get('company_name', 'Construction Co.'))
    ws.append([])
    title_str = f"Diesel Fuel Details Log{' | ' + from_date + ' to ' + to_date if from_date else ''}"
    ws.append([title_str])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color='1A3C5E')
    ws.append([])

    headers = ['I.No', 'Date', 'Vehicle No.', 'Qty (Ltrs)', 'Amount (Rs.)']
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(ws.max_row, col)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
        cell.border = _border()
        cell.alignment = Alignment(horizontal='center')

    even_fill = PatternFill('solid', fgColor='EEF4F8')
    for i, r in enumerate(records):
        ws.append([r['sl_no'], r['date'], r['vehicle_no'], r['qty_liters'], r['amount']])
        fill = even_fill if i % 2 == 1 else None
        for col in range(1, 6):
            cell = ws.cell(ws.max_row, col)
            cell.border = _border()
            if fill:
                cell.fill = fill

    total_qty = sum(r.get('qty_liters', 0) for r in records)
    total_amt = sum(r.get('amount', 0) for r in records)
    ws.append(['', 'TOTAL', '', total_qty, total_amt])
    for col in range(1, 6):
        cell = ws.cell(ws.max_row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D4E8F0')
        cell.border = _border()

    for col_letter, width in [('A', 8), ('B', 14), ('C', 20), ('D', 14), ('E', 16)]:
        ws.column_dimensions[col_letter].width = width

    wb.save(fpath)
    return fpath, None


# ══════════════════════════════════════════════════════════════════════════════
#  CEMENT LOG EXPORTS
# ══════════════════════════════════════════════════════════════════════════════

def export_cement_pdf(records, settings, from_date='', to_date=''):
    if not REPORTLAB_OK:
        return None, "ReportLab not installed. Run: pip install reportlab"
    fname = f"CementLog_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    fpath = os.path.join(get_export_dir(), fname)
    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    s = _pdf_header_style()
    story = []

    story.append(Paragraph(settings.get('company_name', 'Construction Co.'), s['title']))
    story.append(Paragraph('Cement Details Log', s['subtitle']))
    if from_date or to_date:
        story.append(Paragraph(f"Period: {from_date or 'All'} to {to_date or 'All'}", s['subtitle']))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a3c5e')))
    story.append(Spacer(1, 10))

    if records:
        total_qty = sum(r.get('qty', 0) for r in records)
        story.append(Paragraph(
            f"Total Entries: {len(records)}  |  Total Qty: {total_qty:.0f} Bags",
            s['normal']
        ))
        story.append(Spacer(1, 8))

        data = [['Date', 'Qty (Bags)', 'From', 'To', 'Details']]
        for r in records:
            data.append([
                r['date'], f"{r['qty']:.0f}",
                r.get('from_location', '-'), r.get('to_location', '-'),
                r.get('details', '') or ''
            ])
        data.append(['TOTAL', f"{total_qty:.0f}", '', '', ''])

        t = Table(data, colWidths=[3*cm, 3*cm, 4*cm, 4*cm, 5*cm])
        style = _table_style_default()
        style.add('FONTNAME', (0, len(data)-1), (-1, len(data)-1), 'Helvetica-Bold')
        style.add('BACKGROUND', (0, len(data)-1), (-1, len(data)-1), colors.HexColor('#d4e8f0'))
        t.setStyle(style)
        story.append(t)
    else:
        story.append(Paragraph("No cement records found.", s['normal']))

    story.append(Spacer(1, 15))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}", s['subtitle']))
    doc.build(story)
    return fpath, None


def export_cement_excel(records, settings, from_date='', to_date=''):
    if not OPENPYXL_OK:
        return None, "openpyxl not installed. Run: pip install openpyxl"
    fname = f"CementLog_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    fpath = os.path.join(get_export_dir(), fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cement Log'

    _excel_header_style(ws, settings.get('company_name', 'Construction Co.'))
    ws.append([])
    ws.append([f"Cement Details Log{' | ' + from_date + ' to ' + to_date if from_date else ''}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color='1A3C5E')
    ws.append([])

    headers = ['Date', 'Qty (Bags)', 'From', 'To', 'Details']
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(ws.max_row, col)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
        cell.border = _border()
        cell.alignment = Alignment(horizontal='center')

    even_fill = PatternFill('solid', fgColor='EEF4F8')
    for i, r in enumerate(records):
        ws.append([r['date'], r['qty'],
                   r.get('from_location', ''), r.get('to_location', ''),
                   r.get('details', '') or ''])
        fill = even_fill if i % 2 == 1 else None
        for col in range(1, 6):
            cell = ws.cell(ws.max_row, col)
            cell.border = _border()
            if fill:
                cell.fill = fill

    total_qty = sum(r.get('qty', 0) for r in records)
    ws.append(['TOTAL', total_qty, '', '', ''])
    for col in range(1, 6):
        cell = ws.cell(ws.max_row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D4E8F0')
        cell.border = _border()

    for col_letter, width in [('A', 14), ('B', 14), ('C', 25), ('D', 25), ('E', 40)]:
        ws.column_dimensions[col_letter].width = width

    wb.save(fpath)
    return fpath, None
